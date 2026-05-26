"""Report generation: Excel + PDF (Japanese-aware)."""
from __future__ import annotations

import io
from datetime import date
from pathlib import Path

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from src.analytics import (
    get_category_ranking,
    get_channel_breakdown,
    get_data_date_range,
    get_period_aggregates,
    get_vendor_detail,
    get_vendor_ranking,
)
from src.format_utils import format_int

PROJECT_ROOT = Path(__file__).resolve().parents[1]
FONT_DIR = PROJECT_ROOT / "assets" / "fonts"
FONT_GOTHIC = FONT_DIR / "ipaexg.ttf"
FONT_MINCHO = FONT_DIR / "ipaexm.ttf"

_FONTS_REGISTERED = False


def _register_fonts() -> str:
    """Register Japanese fonts. Returns the font name to use."""
    global _FONTS_REGISTERED
    if not _FONTS_REGISTERED:
        if FONT_GOTHIC.exists():
            pdfmetrics.registerFont(TTFont("IPAexGothic", str(FONT_GOTHIC)))
        if FONT_MINCHO.exists():
            pdfmetrics.registerFont(TTFont("IPAexMincho", str(FONT_MINCHO)))
        _FONTS_REGISTERED = True
    return "IPAexGothic" if FONT_GOTHIC.exists() else "Helvetica"


# ---------- Common helpers ----------

def _format_yen(value) -> str:
    import math
    if value is None:
        return "—"
    if isinstance(value, float) and math.isnan(value):
        return "—"
    return f"¥{int(value):,}"


def _format_pct(value) -> str:
    import math
    if value is None:
        return "—"
    if isinstance(value, float) and math.isnan(value):
        return "—"
    return f"{float(value) * 100:.1f}%"


def _resolve_period(
    conn: duckdb.DuckDBPyConnection,
    year: int,
    month: int,
) -> tuple[date, date, str]:
    start = date(year, month, 1)
    if month == 12:
        end = date(year, 12, 31)
    else:
        end = date(year, month + 1, 1) - pd.Timedelta(days=1)
        end = end.date() if hasattr(end, "date") else end
    label = f"{year}年{month:02d}月"
    return start, end, label


# ===========================================================
# Excel report (全社月次)
# ===========================================================

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def _auto_size(ws) -> None:
    for col_cells in ws.columns:
        try:
            length = max(len(str(c.value or "")) for c in col_cells)
        except ValueError:
            length = 10
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 40)


def _write_table(ws, df: pd.DataFrame, start_row: int = 1, start_col: int = 1) -> int:
    if df.empty:
        ws.cell(row=start_row, column=start_col, value="データなし")
        return start_row + 1
    for j, col_name in enumerate(df.columns):
        cell = ws.cell(row=start_row, column=start_col + j, value=str(col_name))
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for i, (_, row) in enumerate(df.iterrows(), start=1):
        for j, value in enumerate(row):
            if pd.isna(value):
                cell_value = ""
            elif isinstance(value, (pd.Timestamp,)):
                cell_value = value.strftime("%Y-%m-%d")
            else:
                cell_value = value
            ws.cell(row=start_row + i, column=start_col + j, value=cell_value)
    return start_row + len(df) + 1


def build_monthly_excel(
    conn: duckdb.DuckDBPyConnection,
    year: int,
    month: int,
    municipality_ids: list[int] | None = None,
) -> bytes:
    start, end, label = _resolve_period(conn, year, month)
    wb = Workbook()

    # --- Sheet 1: サマリー ---
    ws = wb.active
    ws.title = "サマリー"
    ws.cell(row=1, column=1, value=f"HIDAKARAanalytics 月次レポート").font = Font(size=16, bold=True)
    ws.cell(row=2, column=1, value=f"対象月: {label}")
    ws.cell(row=3, column=1, value=f"集計期間: {start} 〜 {end}")

    # KPI section
    muni_clause = ""
    muni_params: list = []
    if municipality_ids:
        placeholders = ",".join(["?"] * len(municipality_ids))
        muni_clause = f" AND municipality_id IN ({placeholders})"
        muni_params = list(municipality_ids)
    row = conn.execute(
        f"""
        SELECT
            COALESCE(SUM(donation_amount), 0) AS revenue,
            COUNT(*) AS orders,
            COALESCE(SUM(product_price), 0) AS total_cost
        FROM shipments
        WHERE payment_date BETWEEN ? AND ? {muni_clause}
        """,
        [start, end] + muni_params,
    ).fetchone()
    revenue, orders, total_cost = int(row[0]), int(row[1]), int(row[2])
    expense_ratio = total_cost / revenue if revenue else 0.0

    kpi_df = pd.DataFrame(
        [
            {"指標": "寄付金額", "値": _format_yen(revenue)},
            {"指標": "件数", "値": f"{orders:,} 件"},
            {"指標": "謝礼品価格", "値": _format_yen(total_cost)},
            {"指標": "経費率", "値": _format_pct(expense_ratio)},
        ]
    )
    _write_table(ws, kpi_df, start_row=5)

    # --- Sheet 2: 月次推移（直近24ヶ月） ---
    ws2 = wb.create_sheet("月次推移")
    trend_df = get_period_aggregates(
        conn,
        start=date(year - 2, month, 1),
        end=end,
        granularity="month",
        municipality_ids=municipality_ids,
    )
    if not trend_df.empty:
        out = trend_df.copy()
        out["月"] = out["period"].dt.strftime("%Y-%m")
        out = out[["月", "orders", "revenue", "total_cost", "expense_ratio", "avg_order_value"]]
        out = out.rename(
            columns={
                "orders": "件数",
                "revenue": "寄付金額",
                "total_cost": "謝礼品価格",
                "expense_ratio": "経費率",
                "avg_order_value": "平均単価",
            }
        )
        _write_table(ws2, out, start_row=1)

        chart = BarChart()
        chart.type = "col"
        chart.style = 11
        chart.title = "月次 寄付金額・謝礼品価格 推移"
        chart.y_axis.title = "金額（円）"
        chart.x_axis.title = "月"
        data = Reference(ws2, min_col=3, min_row=1, max_col=4, max_row=len(out) + 1)
        cats = Reference(ws2, min_col=1, min_row=2, max_row=len(out) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 20
        chart.height = 10
        ws2.add_chart(chart, "H2")

    # --- Sheet 3: カテゴリ別 ---
    ws3 = wb.create_sheet("カテゴリ別")
    cat_df = get_category_ranking(conn, start, end, municipality_ids)
    if not cat_df.empty:
        out = cat_df.copy()
        out = out.rename(
            columns={
                "category": "カテゴリ",
                "orders": "件数",
                "revenue": "寄付金額",
                "total_cost": "謝礼品価格",
                "expense_ratio": "経費率",
                "share": "寄付金額シェア",
            }
        )
        _write_table(ws3, out, start_row=1)

    # --- Sheet 4: 事業者別 ---
    ws4 = wb.create_sheet("事業者別")
    vendor_df = get_vendor_ranking(conn, start, end, municipality_ids=municipality_ids)
    if not vendor_df.empty:
        out = vendor_df.copy()
        out = out.rename(
            columns={
                "vendor": "事業者",
                "orders": "件数",
                "product_count": "取扱商品数",
                "revenue": "寄付金額",
                "total_cost": "謝礼品価格",
                "expense_ratio": "経費率",
                "avg_order_value": "平均単価",
            }
        )
        _write_table(ws4, out, start_row=1)

    # --- Sheet 5: チャネル別 ---
    ws5 = wb.create_sheet("チャネル別")
    ch_df = get_channel_breakdown(conn, start, end, municipality_ids)
    if not ch_df.empty:
        out = ch_df.copy()
        out = out.rename(
            columns={
                "channel": "チャネル",
                "orders": "件数",
                "revenue": "寄付金額",
                "total_cost": "謝礼品価格",
                "expense_ratio": "経費率",
                "share": "寄付金額シェア",
            }
        )
        _write_table(ws5, out, start_row=1)

    for sheet in wb.worksheets:
        _auto_size(sheet)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ===========================================================
# PDF helpers (matplotlib for charts, embedded font)
# ===========================================================

def _setup_matplotlib():
    """Configure matplotlib to use the IPAex Gothic font for Japanese text."""
    import matplotlib

    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib import font_manager

    if FONT_GOTHIC.exists():
        font_manager.fontManager.addfont(str(FONT_GOTHIC))
        matplotlib.rcParams["font.family"] = "IPAexGothic"
    matplotlib.rcParams["axes.unicode_minus"] = False
    return plt


def _chart_to_png_bytes(plt_fig) -> bytes:
    buf = io.BytesIO()
    plt_fig.savefig(buf, format="png", dpi=150, bbox_inches="tight")
    buf.seek(0)
    return buf.getvalue()


def _build_styles(font_name: str):
    base = getSampleStyleSheet()
    styles = {
        "Title": ParagraphStyle(
            "TitleJP", parent=base["Title"], fontName=font_name, fontSize=20, alignment=TA_CENTER, spaceAfter=12
        ),
        "H1": ParagraphStyle(
            "H1JP", parent=base["Heading1"], fontName=font_name, fontSize=14, spaceBefore=12, spaceAfter=6
        ),
        "H2": ParagraphStyle(
            "H2JP", parent=base["Heading2"], fontName=font_name, fontSize=12, spaceBefore=8, spaceAfter=4
        ),
        "Body": ParagraphStyle(
            "BodyJP", parent=base["BodyText"], fontName=font_name, fontSize=10, leading=14, alignment=TA_LEFT
        ),
        "Caption": ParagraphStyle(
            "CaptionJP", parent=base["BodyText"], fontName=font_name, fontSize=8, textColor=colors.grey
        ),
    }
    return styles


def _build_table_style(font_name: str) -> TableStyle:
    return TableStyle(
        [
            ("FONTNAME", (0, 0), (-1, -1), font_name),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FB")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]
    )


def _df_to_table(df: pd.DataFrame, font_name: str, col_widths: list | None = None) -> Table:
    data = [list(df.columns)] + df.astype(str).values.tolist()
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(_build_table_style(font_name))
    return table


# ===========================================================
# PDF report (全社月次)
# ===========================================================

def build_monthly_pdf(
    conn: duckdb.DuckDBPyConnection,
    year: int,
    month: int,
    municipality_ids: list[int] | None = None,
) -> bytes:
    font_name = _register_fonts()
    plt = _setup_matplotlib()
    start, end, label = _resolve_period(conn, year, month)
    styles = _build_styles(font_name)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=f"月次レポート_{label}",
    )

    story = []
    story.append(Paragraph(f"月次レポート {label}", styles["Title"]))
    story.append(Paragraph(f"集計期間: {start} 〜 {end}", styles["Caption"]))
    story.append(Spacer(1, 8))

    # --- KPI summary ---
    muni_clause = ""
    muni_params: list = []
    if municipality_ids:
        placeholders = ",".join(["?"] * len(municipality_ids))
        muni_clause = f" AND municipality_id IN ({placeholders})"
        muni_params = list(municipality_ids)
    row = conn.execute(
        f"""
        SELECT COALESCE(SUM(donation_amount), 0), COUNT(*), COALESCE(SUM(product_price), 0)
        FROM shipments WHERE payment_date BETWEEN ? AND ? {muni_clause}
        """,
        [start, end] + muni_params,
    ).fetchone()
    revenue, orders, total_cost = int(row[0]), int(row[1]), int(row[2])
    expense_ratio = total_cost / revenue if revenue else 0.0

    kpi_df = pd.DataFrame(
        {
            "指標": ["寄付金額", "件数", "謝礼品価格", "経費率"],
            "値": [_format_yen(revenue), f"{orders:,} 件", _format_yen(total_cost), _format_pct(expense_ratio)],
        }
    )
    story.append(Paragraph("◆ KPI サマリー", styles["H1"]))
    story.append(_df_to_table(kpi_df, font_name, col_widths=[60 * mm, 60 * mm]))
    story.append(Spacer(1, 10))

    # --- Monthly trend chart ---
    trend = get_period_aggregates(conn, date(year - 1, month, 1), end, "month", municipality_ids)
    if not trend.empty and len(trend) > 1:
        fig, ax1 = plt.subplots(figsize=(8, 3.5))
        ax2 = ax1.twinx()
        xpos = range(len(trend))
        w = 0.35
        ax1.bar([x - w / 2 for x in xpos], trend["revenue"], width=w, label="寄付金額", color="#1f77b4")
        ax1.bar([x + w / 2 for x in xpos], trend["total_cost"], width=w, label="謝礼品価格", color="#d62728")
        ax2.plot(list(xpos), trend["orders"], "o-", color="#ff7f0e", label="件数")
        ax1.set_xticks(list(xpos))
        ax1.set_xticklabels([d.strftime("%Y/%m") for d in trend["period"]], rotation=30, ha="right")
        ax1.set_ylabel("金額（円）")
        ax2.set_ylabel("件数")
        ax1.legend(loc="upper left")
        ax2.legend(loc="upper right")
        ax1.grid(axis="y", linestyle="--", alpha=0.4)
        story.append(Paragraph("◆ 月次推移", styles["H1"]))
        story.append(Image(io.BytesIO(_chart_to_png_bytes(fig)), width=170 * mm, height=70 * mm))
        plt.close(fig)
        story.append(Spacer(1, 10))

    # --- Category breakdown ---
    cat_df = get_category_ranking(conn, start, end, municipality_ids)
    if not cat_df.empty:
        story.append(Paragraph("◆ カテゴリ別 寄付金額", styles["H1"]))
        d = cat_df.head(15).copy()
        out = pd.DataFrame(
            {
                "カテゴリ": d["category"],
                "件数": d["orders"].map(format_int),
                "寄付金額": d["revenue"].map(_format_yen),
                "謝礼品価格": d["total_cost"].map(_format_yen),
                "経費率": d["expense_ratio"].map(_format_pct),
                "シェア": d["share"].map(_format_pct),
            }
        )
        story.append(_df_to_table(out, font_name, col_widths=[36 * mm, 18 * mm, 32 * mm, 32 * mm, 22 * mm, 22 * mm]))
        story.append(Spacer(1, 10))

    # --- Vendor ranking (top 15) ---
    vendor_df = get_vendor_ranking(conn, start, end, municipality_ids=municipality_ids)
    if not vendor_df.empty:
        story.append(PageBreak())
        story.append(Paragraph("◆ 事業者別ランキング TOP15", styles["H1"]))
        d = vendor_df.head(15).copy()
        out = pd.DataFrame(
            {
                "事業者": d["vendor"].str.slice(0, 28),
                "件数": d["orders"].map(format_int),
                "寄付金額": d["revenue"].map(_format_yen),
                "謝礼品価格": d["total_cost"].map(_format_yen),
                "経費率": d["expense_ratio"].map(_format_pct),
            }
        )
        story.append(_df_to_table(out, font_name, col_widths=[60 * mm, 18 * mm, 35 * mm, 35 * mm, 22 * mm]))
        story.append(Spacer(1, 10))

    # --- Channel breakdown ---
    ch_df = get_channel_breakdown(conn, start, end, municipality_ids)
    if not ch_df.empty:
        story.append(Paragraph("◆ チャネル別 寄付金額構成", styles["H1"]))
        d = ch_df.copy()
        out = pd.DataFrame(
            {
                "チャネル": d["channel"],
                "件数": d["orders"].map(format_int),
                "寄付金額": d["revenue"].map(_format_yen),
                "謝礼品価格": d["total_cost"].map(_format_yen),
                "経費率": d["expense_ratio"].map(_format_pct),
                "シェア": d["share"].map(_format_pct),
            }
        )
        story.append(_df_to_table(out, font_name, col_widths=[36 * mm, 18 * mm, 32 * mm, 32 * mm, 22 * mm, 22 * mm]))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


# ===========================================================
# Vendor-specific PDF report
# ===========================================================

def build_vendor_pdf(
    conn: duckdb.DuckDBPyConnection,
    vendor: str,
    start: date | None = None,
    end: date | None = None,
    municipality_ids: list[int] | None = None,
) -> bytes:
    font_name = _register_fonts()
    plt = _setup_matplotlib()
    styles = _build_styles(font_name)

    min_date, max_date = get_data_date_range(conn, municipality_ids)
    start = start or min_date
    end = end or max_date

    detail = get_vendor_detail(conn, vendor, start, end, municipality_ids=municipality_ids)
    kpi = detail["kpi"]

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=f"事業者レポート_{vendor}",
    )

    story = []
    story.append(Paragraph("事業者別レポート", styles["Title"]))
    story.append(Paragraph(f"事業者名: {vendor}", styles["H1"]))
    story.append(Paragraph(f"集計期間: {start} 〜 {end}", styles["Caption"]))
    story.append(Spacer(1, 6))

    # --- KPI ---
    kpi_df = pd.DataFrame(
        {
            "指標": ["寄付金額", "件数", "謝礼品価格", "経費率", "取扱商品数"],
            "値": [
                _format_yen(kpi.revenue),
                f"{kpi.orders:,} 件",
                _format_yen(kpi.total_cost),
                _format_pct(kpi.expense_ratio),
                f"{detail['product_count']:,}",
            ],
        }
    )
    story.append(Paragraph("◆ サマリー", styles["H1"]))
    story.append(_df_to_table(kpi_df, font_name, col_widths=[60 * mm, 60 * mm]))
    story.append(Spacer(1, 10))

    # --- Monthly trend chart ---
    monthly = detail["monthly"]
    if not monthly.empty and len(monthly) > 1:
        fig, ax = plt.subplots(figsize=(8, 3))
        xpos = range(len(monthly))
        w = 0.35
        ax.bar([x - w / 2 for x in xpos], monthly["revenue"], width=w, label="寄付金額", color="#1f77b4")
        ax.bar([x + w / 2 for x in xpos], monthly["total_cost"], width=w, label="謝礼品価格", color="#d62728")
        ax.set_xticks(list(xpos))
        ax.set_xticklabels([d.strftime("%Y/%m") for d in monthly["month"]], rotation=30, ha="right")
        ax.set_ylabel("金額（円）")
        ax.legend()
        ax.grid(axis="y", linestyle="--", alpha=0.4)
        story.append(Paragraph("◆ 月次推移", styles["H1"]))
        story.append(Image(io.BytesIO(_chart_to_png_bytes(fig)), width=170 * mm, height=60 * mm))
        plt.close(fig)
        story.append(Spacer(1, 10))

    # --- Category / Channel breakdowns ---
    categories = detail["categories"]
    channels = detail["channels"]
    if not categories.empty:
        story.append(Paragraph("◆ カテゴリ別", styles["H1"]))
        d = categories.copy()
        out = pd.DataFrame(
            {
                "カテゴリ": d["category"],
                "件数": d["orders"].map(format_int),
                "寄付金額": d["revenue"].map(_format_yen),
                "謝礼品価格": d["total_cost"].map(_format_yen),
            }
        )
        story.append(_df_to_table(out, font_name, col_widths=[55 * mm, 20 * mm, 45 * mm, 45 * mm]))
        story.append(Spacer(1, 8))

    if not channels.empty:
        story.append(Paragraph("◆ チャネル別", styles["H1"]))
        d = channels.copy()
        out = pd.DataFrame(
            {
                "チャネル": d["channel"],
                "件数": d["orders"].map(format_int),
                "寄付金額": d["revenue"].map(_format_yen),
                "謝礼品価格": d["total_cost"].map(_format_yen),
            }
        )
        story.append(_df_to_table(out, font_name, col_widths=[55 * mm, 20 * mm, 45 * mm, 45 * mm]))
        story.append(Spacer(1, 10))

    # --- Product list (top 30) ---
    products = detail["products"]
    if not products.empty:
        story.append(PageBreak())
        story.append(Paragraph("◆ 商品別実績（TOP30）", styles["H1"]))
        d = products.head(30).copy()
        out = pd.DataFrame(
            {
                "商品名": d["product_name"].str.slice(0, 36),
                "カテゴリ": d["category"].str.slice(0, 10),
                "件数": d["orders"].map(format_int),
                "寄付金額": d["revenue"].map(_format_yen),
                "謝礼品価格": d["total_cost"].map(_format_yen),
                "経費率": d["expense_ratio"].map(_format_pct),
            }
        )
        story.append(
            _df_to_table(
                out,
                font_name,
                col_widths=[60 * mm, 22 * mm, 16 * mm, 30 * mm, 30 * mm, 18 * mm],
            )
        )

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()
